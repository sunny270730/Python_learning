#!/usr/bin/env python
# coding: utf-8

# In[56]:


import openpyxl as xl

def excel_to_dict(sheet, my_dict):                     #  excel name ka function create kiya 
    for row in range(2, sheet.max_row + 1):            # loop ki help se excel sheet check ki
        english_word = sheet.cell(row, 1).value            # 1 row select jis ma english word define ha 

        for column in range(2, sheet.max_column + 1):            # loop ki help se punjabi to english transfer
            if sheet.cell(row, column).value != None:                 # column ma english ja punjabi word nhi output none
                punjabi_word = sheet.cell(row, column).value.replace(" ","")
                my_dict[punjabi_word] = english_word

    return my_dict


def find_in_dict(dictionary, word):
    word = word.replace(" ","")
    try:
        print(dictionary[word])
    except KeyError:
        print('Word not found, Kindly check your input again for mistakes')

my_dict = {}                # dictionary ko call ki
workbook = xl.load_workbook('Test1.xlsx')          # .xlsx file ko access kiya 
sheet = workbook.active
my_sheet = excel_to_dict(sheet, my_dict)
dic = input('enter the punjabi word:- ')
find_in_dict(my_dict,dic)


# In[ ]:




